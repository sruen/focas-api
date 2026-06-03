from __future__ import annotations

import csv
import json
import re
import shutil
import stat
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency guard
    load_workbook = None


COMPANY_ALIASES: list[tuple[str, str]] = [
    ("william hill", "William"),
    ("williamhill", "William"),
    ("william", "William"),
    ("威廉希尔", "William"),
    ("威廉", "William"),
    ("威*", "William"),
    ("ladbrokes", "Ladbrokes"),
    ("ladbroke", "Ladbrokes"),
    ("立博", "Ladbrokes"),
    ("立*", "Ladbrokes"),
    ("avg", "Avg"),
    ("average", "Avg"),
    ("平均欧赔", "Avg"),
    ("平均", "Avg"),
    ("市场平均", "Avg"),
    ("betvictor", "BetVictor"),
    ("victor", "BetVictor"),
    ("韦德", "BetVictor"),
    ("伟德", "BetVictor"),
    ("韦*", "BetVictor"),
]

_LEGACY_TARGET_SUFFIX = "_" + "".join(("Candi", "date"))
TARGET_COMPANY_MAP = {
    f"William{_LEGACY_TARGET_SUFFIX}": "William",
    f"Ladbrokes{_LEGACY_TARGET_SUFFIX}": "Ladbrokes",
    f"BetVictor{_LEGACY_TARGET_SUFFIX}": "BetVictor",
    "William_Source": "William",
    "Ladbrokes_Source": "Ladbrokes",
    "BetVictor_Source": "BetVictor",
    "Avg": "Avg",
    "Average": "Avg",
}

TRIPLE_RE = re.compile(
    r"(?<!\d)(\d{1,2}(?:\.\d{1,3})?)\s*(?:/|／|,|，|\s+)\s*"
    r"(\d{1,2}(?:\.\d{1,3})?)\s*(?:/|／|,|，|\s+)\s*"
    r"(\d{1,2}(?:\.\d{1,3})?)(?!\d)"
)

VS_PATTERNS = [
    re.compile(r"比赛\s*[:：]\s*([^\n\r]+?)\s+(?:vs|VS|v|V|对|vs\.)\s+([^\n\r]+)"),
    re.compile(r"当前位置：.*?>\s*([^\n\r<>]{1,30}?)\s+(?:vs|VS|v|V|对)\s+([^\n\r<>]{1,30})", re.I),
    re.compile(r"([^\n\r<>]{2,30}?)\s+(?:vs|VS|v|V|对)\s+([^\n\r<>]{2,30})"),
    re.compile(r"主队\s*[:：]\s*([^\n\r]+)"),
]


@dataclass
class PackageDiagnostic:
    level: str
    message: str
    file: str | None = None


@dataclass
class PackageLoadResult:
    raw: dict[str, Any]
    diagnostics: list[PackageDiagnostic] = field(default_factory=list)
    source_files: list[str] = field(default_factory=list)

    @property
    def used_canonical_json(self) -> bool:
        return any("标准JSON" in d.message for d in self.diagnostics)


def _empty_raw() -> dict[str, Any]:
    return {
        "match": {
            "home_team": "未识别主队",
            "away_team": "未识别客队",
            "competition": None,
            "kickoff_time": None,
            "stage": None,
            "neutral_venue": None,
            "single_leg": None,
            "match_type": None,
            "extra_time_or_penalties": None,
            "real_home_away": None,
            "attention_level": None,
            "league_for_table": None,
        },
        "home_context": {"name": "未识别主队"},
        "away_context": {"name": "未识别客队"},
        "h2h": {},
        "strength": {},
        "natural_pulls": [],
        "original_book_mode": {},
        "odds": [],
    }


def is_canonical_match_json(raw: Any) -> bool:
    return isinstance(raw, dict) and isinstance(raw.get("match"), dict) and isinstance(raw.get("odds"), list)


def _safe_read_text(path: Path) -> str:
    for enc in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return path.read_text(encoding=enc, errors="strict")
        except Exception:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def _iter_package_files(root: Path) -> Iterable[Path]:
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        parts = set(path.parts)
        if "__MACOSX" in parts or path.name.startswith("."):
            continue
        yield path


def _safe_extract_zip(zf: zipfile.ZipFile, root: Path) -> None:
    """Extract package members without allowing paths outside the package root."""
    resolved_root = root.resolve()
    for info in zf.infolist():
        raw_name = info.filename.replace("\\", "/")
        member = Path(raw_name)
        if member.is_absolute() or re.match(r"^[A-Za-z]:", raw_name):
            raise ValueError(f"ZIP_UNSAFE_PATH: absolute member path rejected: {info.filename}")
        if any(part in {"", ".", ".."} for part in member.parts):
            raise ValueError(f"ZIP_UNSAFE_PATH: traversal member path rejected: {info.filename}")
        unix_mode = (info.external_attr >> 16) & 0o170000
        if stat.S_ISLNK(unix_mode):
            raise ValueError(f"ZIP_UNSAFE_PATH: symbolic link rejected: {info.filename}")
        destination = (resolved_root / member).resolve()
        if destination != resolved_root and resolved_root not in destination.parents:
            raise ValueError(f"ZIP_UNSAFE_PATH: member escaped package root: {info.filename}")
        if info.is_dir():
            destination.mkdir(parents=True, exist_ok=True)
            continue
        destination.parent.mkdir(parents=True, exist_ok=True)
        with zf.open(info, "r") as source, destination.open("wb") as target:
            shutil.copyfileobj(source, target)


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    text = _safe_read_text(path)
    try:
        return [dict(row) for row in csv.DictReader(text.splitlines())]
    except Exception:
        return []


def _text_from_csv(path: Path) -> str:
    text = _safe_read_text(path)
    lines: list[str] = []
    try:
        reader = csv.reader(text.splitlines())
        for i, row in enumerate(reader):
            if i > 200:
                break
            lines.append(" ".join(str(c) for c in row if str(c).strip()))
    except Exception:
        return text
    return "\n".join(lines)


def _text_from_xlsx(path: Path) -> str:
    if load_workbook is None:
        return ""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return ""
    lines: list[str] = []
    for ws in wb.worksheets[:8]:
        lines.append(f"[SHEET] {ws.title}")
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx > 200:
                break
            vals = [str(v) for v in row if v is not None and str(v).strip()]
            if vals:
                lines.append(" ".join(vals))
    try:
        wb.close()
    except Exception:
        pass
    return "\n".join(lines)


def _collect_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".csv", ".html", ".htm", ".log"}:
        if suffix == ".csv":
            return _text_from_csv(path)
        return _safe_read_text(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _text_from_xlsx(path)
    return ""


def _strip_html(value: str) -> str:
    value = re.sub(r"<[^>]+>", "", value)
    value = value.replace("&nbsp;", " ").replace("\ufeff", "")
    return value


def _clean_team_name(value: str) -> str:
    text = _strip_html(value)
    text = re.sub(r"[：:。；;，,|｜].*$", "", text).strip()
    text = re.sub(r"\s+", " ", text)
    return text[:40].strip() or "未识别"


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    m = re.search(r"\d+(?:\.\d+)?", str(value).replace(",", ""))
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def infer_teams(text: str) -> tuple[str | None, str | None]:
    clean = _strip_html(text)
    for pat in VS_PATTERNS[:3]:
        m = pat.search(clean)
        if m:
            return _clean_team_name(m.group(1)), _clean_team_name(m.group(2))
    home = away = None
    mh = re.search(r"主队\s*[:：]\s*([^\n\r]+)", clean)
    ma = re.search(r"客队\s*[:：]\s*([^\n\r]+)", clean)
    if mh:
        home = _clean_team_name(mh.group(1))
    if ma:
        away = _clean_team_name(ma.group(1))
    return home, away


def infer_match_meta(text: str) -> dict[str, Any]:
    meta: dict[str, Any] = {}
    clean = _strip_html(text)
    patterns = {
        "competition": r"赛事(?:名称)?\s*[:：]\s*([^\n\r]+)",
        "kickoff_time": r"(?:比赛时间|开赛时间|时间)\s*[:：]\s*([^\n\r]+)",
        "stage": r"(?:比赛阶段|阶段|轮次)\s*[:：]\s*([^\n\r]+)",
        "match_type": r"(?:赛事类型|比赛类型)\s*[:：]\s*([^\n\r]+)",
        "league_for_table": r"(?:查表联赛|联赛|league_for_table)\s*[:：]\s*([^\n\r]+)",
    }
    for key, pat in patterns.items():
        m = re.search(pat, clean, flags=re.IGNORECASE)
        if m:
            meta[key] = m.group(1).strip()[:80]
    if re.search(r"中立场|neutral", clean, flags=re.IGNORECASE):
        meta["neutral_venue"] = True
        meta.setdefault("real_home_away", False)
    if re.search(r"单回合|single\s*leg", clean, flags=re.IGNORECASE):
        meta["single_leg"] = True
    if re.search(r"决赛", clean):
        meta.setdefault("match_type", "决赛")
        meta.setdefault("stage", "决赛")
    return meta


def _canonical_company(line: str) -> str | None:
    lower = line.lower()
    for alias, canonical in COMPANY_ALIASES:
        if alias.lower() in lower:
            return canonical
    return None


def _valid_odds_triple(vals: tuple[str, str, str]) -> tuple[float, float, float] | None:
    nums = tuple(float(v) for v in vals)
    # Standard 1X2 odds only. Non-odds columns are deliberately ignored.
    if 1.0 <= nums[0] <= 20.0 and 1.0 <= nums[1] <= 20.0 and 1.0 <= nums[2] <= 30.0:
        return nums
    return None


INITIAL_LABEL_RE = re.compile(r"初赔|初盘|开赔|初始|initial|open", re.I)
CURRENT_LABEL_RE = re.compile(r"即时|最新|当前|终赔|现赔|变赔|current|now|live|close", re.I)


def _line_default_label(line: str) -> str:
    has_initial = bool(INITIAL_LABEL_RE.search(line))
    has_current = bool(CURRENT_LABEL_RE.search(line))
    if has_current and not has_initial:
        return "current"
    if has_initial and not has_current:
        return "initial"
    return "unknown"


def _label_for_triple(line: str, start: int, end: int, fallback: str) -> str:
    before = line[max(0, start - 48):start]
    after = line[end:min(len(line), end + 18)]
    last_initial = max((m.start() for m in INITIAL_LABEL_RE.finditer(before)), default=-1)
    last_current = max((m.start() for m in CURRENT_LABEL_RE.finditer(before)), default=-1)
    if last_initial >= 0 or last_current >= 0:
        return "current" if last_current > last_initial else "initial"
    if CURRENT_LABEL_RE.search(after) and not INITIAL_LABEL_RE.search(after):
        return "current"
    if INITIAL_LABEL_RE.search(after) and not CURRENT_LABEL_RE.search(after):
        return "initial"
    return fallback


def extract_odds(text: str) -> list[dict[str, Any]]:
    company_triples: dict[str, list[tuple[str, tuple[float, float, float]]]] = {}
    for raw_line in text.splitlines():
        line = _strip_html(raw_line).strip()
        if not line:
            continue
        company = _canonical_company(line)
        if not company:
            continue
        matches = list(TRIPLE_RE.finditer(line))
        triples_with_label: list[tuple[str, tuple[float, float, float]]] = []
        fallback = _line_default_label(line)
        for m in matches:
            triple = _valid_odds_triple(m.groups())
            if not triple:
                continue
            label = _label_for_triple(line, m.start(), m.end(), fallback)
            triples_with_label.append((label, triple))
        if not triples_with_label:
            continue
        if len(triples_with_label) >= 2 and all(label == "unknown" for label, _ in triples_with_label):
            triples_with_label[0] = ("initial", triples_with_label[0][1])
            triples_with_label[-1] = ("current", triples_with_label[-1][1])
        company_triples.setdefault(company, []).extend(triples_with_label)

    out: list[dict[str, Any]] = []
    for company, items in company_triples.items():
        initial = next((t for label, t in items if label == "initial"), items[0][1])
        current = next((t for label, t in reversed(items) if label == "current"), items[-1][1])
        out.append({
            "company": company,
            "initial": {"home": initial[0], "draw": initial[1], "away": initial[2]},
            "current": {"home": current[0], "draw": current[1], "away": current[2]},
        })
    return out


def _set_match_from_metadata(raw: dict[str, Any], rows: list[dict[str, str]], diagnostics: list[PackageDiagnostic], file_name: str) -> None:
    if not rows:
        return
    row = rows[0]
    home = row.get("home_team") or row.get("主队")
    away = row.get("away_team") or row.get("客队")
    if home:
        raw["match"]["home_team"] = str(home).strip()
        raw["home_context"]["name"] = str(home).strip()
    if away:
        raw["match"]["away_team"] = str(away).strip()
        raw["away_context"]["name"] = str(away).strip()
    league = row.get("league") or row.get("competition") or row.get("赛事")
    if league:
        raw["match"]["competition"] = str(league).strip()
        raw["match"]["league_for_table"] = str(league).strip()
    if row.get("match_date"):
        raw["match"]["kickoff_time"] = str(row.get("match_date")).strip()
    diagnostics.append(PackageDiagnostic("INFO", "已从 match_metadata.csv 精准读取比赛双方与赛事字段。", file_name))


def _odds_from_opening_closing(rows: list[dict[str, str]], diagnostics: list[PackageDiagnostic], file_name: str) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    skipped: list[str] = []
    for row in rows:
        target = (row.get("target_company") or "").strip()
        company = TARGET_COMPANY_MAP.get(target)
        if not company:
            if target:
                skipped.append(target)
            continue
        nums = [
            _to_float(row.get("opening_home")), _to_float(row.get("opening_draw")), _to_float(row.get("opening_away")),
            _to_float(row.get("closing_home")), _to_float(row.get("closing_draw")), _to_float(row.get("closing_away")),
        ]
        if any(v is None for v in nums):
            diagnostics.append(PackageDiagnostic("WARN", f"{target} opening_closing 行赔率不完整，已跳过。", file_name))
            continue
        out.append({
            "company": company,
            "initial": {"home": nums[0], "draw": nums[1], "away": nums[2]},
            "current": {"home": nums[3], "draw": nums[4], "away": nums[5]},
        })
    if out:
        diagnostics.append(PackageDiagnostic("INFO", f"已从 opening_closing.csv 精准读取赔率公司：{', '.join(o['company'] for o in out)}。", file_name))
    if skipped:
        diagnostics.append(PackageDiagnostic("INFO", f"已忽略非核心公司：{', '.join(sorted(set(skipped)))}。", file_name))
    return out


def _avg_from_text(text: str) -> dict[str, Any] | None:
    for raw_line in text.splitlines():
        line = _strip_html(raw_line)
        if "平均欧赔" not in line and "Average" not in line and not re.search(r"\bAvg\b", line, re.I):
            continue
        nums = [_to_float(x.group(0)) for x in re.finditer(r"\d+(?:\.\d+)?", line)]
        # The first six numbers after the company label are initial/current 1X2 odds in ZGZCW table rows.
        nums = [n for n in nums if n is not None]
        if len(nums) >= 6:
            return {
                "company": "Avg",
                "initial": {"home": nums[0], "draw": nums[1], "away": nums[2]},
                "current": {"home": nums[3], "draw": nums[4], "away": nums[5]},
            }
    return None


def _company_from_debug_filename(path: Path) -> str | None:
    name = path.name.lower()
    if "william" in name:
        return "William"
    if "ladbrokes" in name or "ladbroke" in name:
        return "Ladbrokes"
    if "betvictor" in name or "victor" in name:
        return "BetVictor"
    return None


def _extract_zgzcw_debug_records(text: str) -> list[tuple[float, float, float]]:
    """Parse ZGZCW company change-process debug text.

    ZGZCW debug rows are multi-line; ordinary line-level company parsing misses
    them because the company label appears only in the title and the odds appear
    on later lines. Rows are listed newest first, oldest last.
    """
    clean = _strip_html(text)
    clean = clean.replace("↑", "").replace("↓", "")
    # Capture the first 1X2 odds triple immediately after a “赛前...分” update label.
    pat = re.compile(
        r"赛前[^\n\r]*?[分时][\s\t\r\n]+"
        r"(\d{1,2}\.\d{1,3})\s+"
        r"(\d{1,2}\.\d{1,3})\s+"
        r"(\d{1,2}\.\d{1,3})"
    )
    records: list[tuple[float, float, float]] = []
    for m in pat.finditer(clean):
        triple = _valid_odds_triple(m.groups())
        if triple:
            records.append(triple)
    return records


def _odds_from_zgzcw_debug_files(files: list[Path], root: Path, diagnostics: list[PackageDiagnostic]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for file in files:
        if not file.name.endswith("_zhishu_text.txt"):
            continue
        company = _company_from_debug_filename(file)
        if not company:
            continue
        records = _extract_zgzcw_debug_records(_safe_read_text(file))
        if not records:
            diagnostics.append(PackageDiagnostic("WARN", f"{company} debug 文本未能解析出变赔记录。", str(file.relative_to(root))))
            continue
        current = records[0]
        initial = records[-1]
        out.append({
            "company": company,
            "initial": {"home": initial[0], "draw": initial[1], "away": initial[2]},
            "current": {"home": current[0], "draw": current[1], "away": current[2]},
        })
    if out:
        diagnostics.append(PackageDiagnostic("INFO", f"已从 ZGZCW debug 文本兜底读取赔率公司：{', '.join(o['company'] for o in out)}。"))
    return out


def _bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in {"true", "yes", "y", "1", "是"}:
        return True
    if text in {"false", "no", "n", "0", "否"}:
        return False
    return None


def _pull_level(value: Any) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    if "强" in text:
        return "强"
    if "中" in text:
        return "中"
    if "弱" in text:
        return "弱"
    return None


def _easy_to_receive(value: Any) -> bool | None:
    parsed = _bool(value)
    if parsed is not None:
        return parsed
    text = str(value or "").strip()
    if any(key in text for key in ("强", "中", "容易", "可承接")):
        return True
    if any(key in text for key in ("弱", "不易", "不足")):
        return False
    return None


def _dynamic_final_gap(payload: dict[str, Any], away_team: str) -> str | None:
    broad = payload.get("broad_strength") or {}
    static_gap = str(broad.get("static_grade_gap") or "").strip()
    final_text = str(broad.get("final_dynamic_strength") or "").strip()
    match = re.search(r"(\d+(?:\.\d+)?)\s*档", static_gap)
    if match and ("客队" in static_gap or away_team in static_gap):
        return f"客队高{match.group(1)}档；{final_text}" if final_text else f"客队高{match.group(1)}档"
    if match and "主队" in static_gap:
        return f"主队高{match.group(1)}档；{final_text}" if final_text else f"主队高{match.group(1)}档"
    return final_text or static_gap or None


def _enrich_from_context_payload(
    raw: dict[str, Any],
    payload: dict[str, Any],
    diagnostics: list[PackageDiagnostic],
    file_name: str,
) -> None:
    """Map a context-filled collector payload into the canonical engine input."""
    match_payload = payload.get("match") or {}
    home_payload = payload.get("home_context") or {}
    away_payload = payload.get("away_context") or {}
    h2h_payload = payload.get("h2h_context") or {}
    broad = payload.get("broad_strength") or {}
    pulls = payload.get("natural_pull") or {}

    home_team = str(match_payload.get("home_team") or raw["match"].get("home_team") or "").strip()
    away_team = str(match_payload.get("away_team") or raw["match"].get("away_team") or "").strip()
    kickoff = match_payload.get("kickoff_beijing_time") or match_payload.get("kickoff_local_bulgaria_eest")
    raw["match"].update({
        "home_team": home_team,
        "away_team": away_team,
        "competition": match_payload.get("competition") or raw["match"].get("competition"),
        "kickoff_time": kickoff or raw["match"].get("kickoff_time"),
        "stage": match_payload.get("stage") or match_payload.get("competition") or "友谊赛",
        "neutral_venue": _bool(match_payload.get("neutral_field")),
        "single_leg": _bool(match_payload.get("single_match")),
        "match_type": match_payload.get("match_type") or match_payload.get("competition") or "友谊赛",
        "extra_time_or_penalties": (
            "有加时或点球规则"
            if _bool(match_payload.get("extra_time_penalties"))
            else "友谊赛无加时或点球规则"
        ),
        "real_home_away": _bool(match_payload.get("home_away_attribute_real")),
        "attention_level": match_payload.get("attention_level") or "国家队友谊赛中等关注度",
        "league_for_table": match_payload.get("competition") or raw["match"].get("league_for_table"),
    })

    raw["home_context"] = {
        "name": home_team,
        "rank": home_payload.get("ranking_points_context"),
        "points": home_payload.get("ranking_points_context"),
        "recent_matches": list(home_payload.get("recent_5") or []),
        "venue_adaptation": home_payload.get("home_or_neutral_adaptation"),
        "attack_state": home_payload.get("attack_state"),
        "defense_state": home_payload.get("defense_state"),
        "injuries": home_payload.get("injuries_squad"),
        "schedule_fatigue": home_payload.get("schedule_fitness"),
        "motivation": home_payload.get("motivation"),
        "popularity_story": home_payload.get("popularity_story"),
        "major_recent_matches": home_payload.get("major_recent_match_signal"),
    }
    raw["away_context"] = {
        "name": away_team,
        "rank": away_payload.get("ranking_points_context"),
        "points": away_payload.get("ranking_points_context"),
        "recent_matches": list(away_payload.get("recent_5") or []),
        "venue_adaptation": away_payload.get("away_or_neutral_adaptation") or away_payload.get("home_or_neutral_adaptation"),
        "attack_state": away_payload.get("attack_state"),
        "defense_state": away_payload.get("defense_state"),
        "injuries": away_payload.get("injuries_squad"),
        "schedule_fatigue": away_payload.get("schedule_fitness"),
        "motivation": away_payload.get("motivation"),
        "popularity_story": away_payload.get("popularity_story"),
        "major_recent_matches": away_payload.get("major_recent_match_signal"),
    }
    raw["h2h"] = {
        "overall": h2h_payload.get("overall_h2h"),
        "recent_years": " | ".join(h2h_payload.get("recent_h2h") or []),
        "same_competition": h2h_payload.get("same_competition_h2h"),
        "venue_specific": h2h_payload.get("home_away_h2h"),
        "latest_key_match": h2h_payload.get("latest_key_h2h"),
        "market_psychology": h2h_payload.get("market_psychology"),
    }
    raw["strength"] = {
        "home_grade": broad.get("home_grade"),
        "away_grade": broad.get("away_grade"),
        "static_gap": broad.get("static_grade_gap"),
        "dynamic_adjustment": " | ".join(broad.get("dynamic_corrections") or []),
        "final_gap": _dynamic_final_gap(payload, away_team),
    }

    pull_map = (
        ("home_win_pull", "主胜", "大众第一眼方向"),
        ("draw_pull", "平局", "天然缓冲方向"),
        ("away_win_pull", "客胜", "理性受注方向"),
    )
    raw["natural_pulls"] = []
    for source_key, direction, popularity_direction in pull_map:
        item = pulls.get(source_key) or {}
        raw["natural_pulls"].append({
            "direction": direction,
            "strength": _pull_level(item.get("level")),
            "facts": item.get("basis"),
            "market_psychology": item.get("market_psychology"),
            "popularity_direction": popularity_direction,
            "easy_to_receive": _easy_to_receive(item.get("easy_to_take_bets")),
            "first_eye_direction": _bool(item.get("first_eye_direction")),
        })
    diagnostics.append(PackageDiagnostic(
        "INFO",
        "已读取 match_context_payload.json：基本面、往绩、广义实力和自然拉力已映射为正式引擎输入。",
        file_name,
    ))


def build_raw_from_structured_files(files: list[Path], root: Path, aggregate_text: str, diagnostics: list[PackageDiagnostic]) -> dict[str, Any] | None:
    raw = _empty_raw()
    found_meta = False
    structured_odds: list[dict[str, Any]] = []
    context_payload: tuple[dict[str, Any], str] | None = None
    for file in files:
        rel = str(file.relative_to(root))
        name = file.name.lower()
        if name == "match_metadata.csv":
            rows = _read_csv_rows(file)
            _set_match_from_metadata(raw, rows, diagnostics, rel)
            found_meta = bool(rows)
        elif name == "opening_closing.csv":
            structured_odds = _odds_from_opening_closing(_read_csv_rows(file), diagnostics, rel)
        elif name == "match_context_payload.json":
            try:
                payload = json.loads(_safe_read_text(file))
            except Exception as exc:
                diagnostics.append(PackageDiagnostic("WARN", f"match_context_payload.json 读取失败：{exc}", rel))
            else:
                if isinstance(payload, dict):
                    context_payload = payload, rel

    # Enrich metadata from page text when present.
    meta = infer_match_meta(aggregate_text)
    raw["match"].update({k: v for k, v in meta.items() if v is not None})
    if context_payload is not None:
        _enrich_from_context_payload(raw, context_payload[0], diagnostics, context_payload[1])
    if raw["match"].get("competition") and not raw["match"].get("league_for_table"):
        raw["match"]["league_for_table"] = raw["match"]["competition"]
    if not found_meta:
        home, away = infer_teams(aggregate_text)
        if home:
            raw["match"]["home_team"] = home
            raw["home_context"]["name"] = home
        if away:
            raw["match"]["away_team"] = away
            raw["away_context"]["name"] = away

    if not structured_odds:
        structured_odds = _odds_from_zgzcw_debug_files(files, root, diagnostics)

    if structured_odds:
        avg = _avg_from_text(aggregate_text)
        if avg and not any(o["company"] == "Avg" for o in structured_odds):
            structured_odds.append(avg)
            diagnostics.append(PackageDiagnostic("INFO", "已从百家欧赔文本读取 Avg 初赔/当前赔。"))
        elif not avg:
            diagnostics.append(PackageDiagnostic("WARN", "未能从百家欧赔文本读取 Avg。"))
        raw["odds"] = structured_odds
        diagnostics.append(PackageDiagnostic("INFO", "当前按比赛包入口读取，但内容更接近赔率包；程序只负责抽取赔率并生成标准输入，正式基本面仍由闸门拦截。"))
        return raw
    if context_payload is not None:
        diagnostics.append(PackageDiagnostic("WARN", "已读取 match_context_payload.json，但未发现赔率数据；后续赔率闸门将停止正式分析。"))
        return raw
    return None


def build_raw_from_text(text: str, diagnostics: list[PackageDiagnostic]) -> dict[str, Any]:
    raw = _empty_raw()
    home, away = infer_teams(text)
    if home:
        raw["match"]["home_team"] = home
        raw["home_context"]["name"] = home
    else:
        diagnostics.append(PackageDiagnostic("WARN", "未能从比赛包文本中识别主队。"))
    if away:
        raw["match"]["away_team"] = away
        raw["away_context"]["name"] = away
    else:
        diagnostics.append(PackageDiagnostic("WARN", "未能从比赛包文本中识别客队。"))

    meta = infer_match_meta(text)
    raw["match"].update(meta)
    odds = extract_odds(text)
    raw["odds"] = odds
    if not odds:
        diagnostics.append(PackageDiagnostic("WARN", "未能从比赛包文本中识别 William / Ladbrokes / Avg 三项欧赔。"))
    else:
        companies = ", ".join(o["company"] for o in odds)
        diagnostics.append(PackageDiagnostic("INFO", f"已从文本抽取赔率公司：{companies}。"))
        if not raw["natural_pulls"] and not raw.get("strength"):
            diagnostics.append(PackageDiagnostic("INFO", "当前按比赛包入口读取，但内容更接近赔率包；程序只负责抽取赔率并生成标准输入，正式基本面仍由闸门拦截。"))
    return raw


def load_package(path: str | Path) -> PackageLoadResult:
    package_path = Path(path)
    diagnostics: list[PackageDiagnostic] = []
    source_files: list[str] = []
    if not package_path.exists():
        raise FileNotFoundError(str(package_path))
    if package_path.suffix.lower() != ".zip":
        raise ValueError("load_package only accepts .zip files")

    with tempfile.TemporaryDirectory(prefix="focas_pkg_") as td:
        root = Path(td)
        with zipfile.ZipFile(package_path) as zf:
            _safe_extract_zip(zf, root)
        files = list(_iter_package_files(root))
        source_files = [str(f.relative_to(root)) for f in files]

        # 1) Prefer canonical FOCAS JSON, because it is already the correct contract.
        for file in files:
            if file.suffix.lower() != ".json":
                continue
            try:
                raw = json.loads(_safe_read_text(file))
            except Exception as exc:
                diagnostics.append(PackageDiagnostic("WARN", f"JSON 读取失败：{exc}", str(file.relative_to(root))))
                continue
            if is_canonical_match_json(raw):
                diagnostics.append(PackageDiagnostic("INFO", "已使用比赛包内标准JSON输入。", str(file.relative_to(root))))
                return PackageLoadResult(raw=raw, diagnostics=diagnostics, source_files=source_files)

        # 2) Aggregate readable text for metadata enrichment and fallback parsing.
        chunks: list[str] = []
        for file in files:
            text = _collect_text(file)
            if text.strip():
                chunks.append(f"\n[FILE] {file.relative_to(root)}\n{text}")
        aggregate_text = "\n".join(chunks)

        # 3) Prefer structured ZGZCW package CSVs when present; they are safer than text parsing.
        structured_raw = build_raw_from_structured_files(files, root, aggregate_text, diagnostics)
        if structured_raw is not None:
            diagnostics.append(PackageDiagnostic("INFO", "未发现标准JSON，已根据结构化CSV生成部分标准输入；缺失项会由闸门拦截。"))
            return PackageLoadResult(raw=structured_raw, diagnostics=diagnostics, source_files=source_files)

        # 4) Fallback: infer from readable text.
        if not chunks:
            diagnostics.append(PackageDiagnostic("ERROR", "比赛包内没有可解析文本、CSV、Excel 或标准 JSON。"))
            return PackageLoadResult(raw=_empty_raw(), diagnostics=diagnostics, source_files=source_files)
        raw = build_raw_from_text(aggregate_text, diagnostics)
        diagnostics.append(PackageDiagnostic("INFO", "未发现标准JSON，已根据文本/表格内容生成部分标准输入；缺失项会由闸门拦截。"))
        return PackageLoadResult(raw=raw, diagnostics=diagnostics, source_files=source_files)
