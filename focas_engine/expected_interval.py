from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any, Optional

import openpyxl

from .models import ExpectedOpeningInterval, IntervalAuditResult, OpeningIntervalAudit, SkeletonIntervalProfile, StrengthContext
from .table_lookup import load_interval_profile

HARD_COMPANIES = {"威廉", "立博", "William", "Ladbrokes"}
LOW_SIDE_TO_DIRECTION = {"主低赔": "主胜", "客低赔": "客胜", "平低赔_特殊": "平局"}
COMPANY_ALIAS = {"William": "威廉", "Ladbrokes": "立博", "Avg": "市场平均"}
P4_BRIDGE_SHEET = "focas_engine/data/p4_strength_interval_table.csv"
DEFAULT_P4_BRIDGE = Path(__file__).parent / "data" / "p4_strength_interval_table.csv"
V6_GRADE_MAP_SHEET = "\u6863\u6b21\u6620\u5c04"
V6_GRADE_ORDER = [
    "\u4e0b\u6e38",
    "\u4e2d\u4e0b",
    "\u4e2d\u6e38",
    "\u4e2d\u4e0a",
    "\u4e2d\u5f3a",
    "\u51c6\u5f3a",
    "\u666e\u5f3a",
    "\u4eba\u5f3a",
]
V6_GRADE_SCORE = {grade: index for index, grade in enumerate(V6_GRADE_ORDER)}


def _company_cn(name: str) -> str:
    return COMPANY_ALIAS.get(name, name)


def _cn_number_to_float(text: str) -> Optional[float]:
    if not text:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)\s*档", text)
    if match:
        return float(match.group(1))
    if "半档" in text:
        return 0.5
    if "一档" in text:
        return 1.0
    if "两档" in text or "二档" in text:
        return 2.0
    if "三档" in text:
        return 3.0
    return None


def parse_gap_value(strength: StrengthContext, estimate=None) -> Optional[float]:
    """Return the home-minus-away dynamic broad-strength relation."""
    text = " ".join(str(x) for x in (strength.final_gap, strength.static_gap) if x)
    if text:
        amount = _cn_number_to_float(text)
        if amount is not None:
            if "客" in text and ("高" in text or "强" in text or "优势" in text):
                return -amount
            if "主" in text and ("高" in text or "强" in text or "优势" in text):
                return amount
        if "同档" in text or "近似同档" in text or "均势" in text:
            return 0.0

    value = getattr(estimate, "final_gap_value", None)
    if value is not None:
        try:
            return float(value)
        except Exception:
            return None
    return None


def _gap_label(value: Optional[float]) -> str:
    if value is None:
        return "未解析"
    if abs(value) < 0.25:
        return "双方同档或近似同档"
    side = "主队" if value > 0 else "客队"
    amount = abs(value)
    amount_text = "一" if amount == 1.0 else "两" if amount == 2.0 else f"{amount:g}"
    return f"{side}高{amount_text}档"


def p4_strength_key(gap: Optional[float]) -> Optional[str]:
    if gap is None:
        return None
    return f"gap:{round(gap * 2) / 2:.1f}"


def _normalize_grade(value: Any) -> Optional[str]:
    text = re.sub(r"\s+", "", str(value or ""))
    if not text:
        return None
    for grade in V6_GRADE_ORDER:
        if grade in text:
            return grade
    return None


def _parse_v6_zone_cell(value: Any) -> tuple[Optional[int], Optional[int], str, str]:
    text = str(value or "").strip()
    if not text:
        return None, None, "", "EMPTY"
    if "\u9006" in text:
        return 0, 0, text, "REVERSE_MAP_REVIEW"
    match = re.search(r"(\d+)\s*(?:[-~\u2014]\s*(\d+))?", text)
    if not match:
        return None, None, text, "UNPARSED"
    lower = int(match.group(1))
    upper = int(match.group(2)) if match.group(2) else lower
    if upper != lower:
        return lower, upper, text, "RANGE_INTERVAL"
    if "+" in text:
        return lower, 12, text, "PLUS_INTERVAL"
    return lower, upper, text, "EXACT"


def _v6_low_side(home_grade: str, away_grade: str) -> str:
    home_score = V6_GRADE_SCORE.get(home_grade)
    away_score = V6_GRADE_SCORE.get(away_grade)
    if home_score is None or away_score is None:
        return "\u672a\u89e3\u6790"
    return "\u4e3b\u4f4e\u8d54" if home_score >= away_score else "\u5ba2\u4f4e\u8d54"


def _read_v6_grade_mapping(table_path: str | Path | None) -> list[dict[str, Any]]:
    if table_path is None:
        return []
    path = Path(table_path)
    if not path.is_file():
        return []
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if V6_GRADE_MAP_SHEET not in workbook.sheetnames:
            return []
        sheet = workbook[V6_GRADE_MAP_SHEET]
        rows = list(sheet.iter_rows(values_only=True))
        header_index: Optional[int] = None
        away_headers: list[Optional[str]] = []
        for index, values in enumerate(rows):
            normalized = [_normalize_grade(value) for value in values]
            if sum(1 for value in normalized if value in V6_GRADE_SCORE) >= 4:
                header_index = index
                away_headers = normalized
                break
        if header_index is None:
            return []
        mapped: list[dict[str, Any]] = []
        for row_number, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
            home_grade = _normalize_grade(values[0] if values else None)
            if home_grade not in V6_GRADE_SCORE:
                continue
            for col_index, away_grade in enumerate(away_headers):
                if col_index == 0 or away_grade not in V6_GRADE_SCORE:
                    continue
                cell = values[col_index] if col_index < len(values) else None
                interval_min, interval_max, raw_zone, parse_status = _parse_v6_zone_cell(cell)
                if interval_min is None or interval_max is None:
                    continue
                mapped.append(
                    {
                        "home_grade": home_grade,
                        "away_grade": away_grade,
                        "interval": interval_min,
                        "interval_min": interval_min,
                        "interval_max": interval_max,
                        "raw_zone": raw_zone,
                        "parse_status": parse_status,
                        "row_number": row_number,
                        "low_side": _v6_low_side(home_grade, away_grade),
                    }
                )
        return mapped
    finally:
        workbook.close()


def _expected_interval_from_v6_grade_map(
    *,
    table_path: str | Path | None,
    strength: StrengthContext,
    gap: Optional[float],
) -> ExpectedOpeningInterval | None:
    home_grade = _normalize_grade(strength.home_grade)
    away_grade = _normalize_grade(strength.away_grade)
    if not home_grade or not away_grade:
        return None
    rows = _read_v6_grade_mapping(table_path)
    for row in rows:
        if row["home_grade"] != home_grade or row["away_grade"] != away_grade:
            continue
        raw_zone = str(row["raw_zone"])
        interval_min = int(row["interval_min"])
        interval_max = int(row["interval_max"])
        callable_ids = list(range(interval_min, interval_max + 1))
        notes = [
            f"V6_GRADE_MAP: home_grade={home_grade}; away_grade={away_grade}; raw_zone={raw_zone}.",
            "V6 uses the low-odds skeleton axis. Draw and opposite-win columns are market-ladder references only.",
        ]
        if row["parse_status"] in {"RANGE_INTERVAL", "PLUS_INTERVAL", "REVERSE_MAP_REVIEW"}:
            notes.append(
                f"V6_GRADE_MAP_REVIEW: {row['parse_status']}; callable_interval_ids={callable_ids}; do not collapse the raw zone into one fixed interval."
            )
        return _expected(
            gap=gap,
            low_side=row["low_side"],
            interval=interval_min,
            water="\u8868\u5185\u6c34\u4f4d\u5f85\u5ba1\u8ba1",
            source="STRENGTH_INTERVAL_BRIDGE",
            confidence="TABLE_CONFIRMED" if row["parse_status"] == "EXACT" else "TABLE_RANGE_REVIEW",
            rule="V6 grade mapping sheet routes home/away broad-strength grades to the modern low-odds skeleton interval.",
            notes=notes,
            strength=strength,
            matched_row_id=f"V6:{home_grade}:{away_grade}:{raw_zone}",
            matched_row_number=int(row["row_number"]),
            lookup_key_status="V6_GRADE_MAP_MATCHED",
            matched_sheet=f"{Path(table_path).name}::{V6_GRADE_MAP_SHEET}" if table_path else V6_GRADE_MAP_SHEET,
            interval_min=interval_min,
            interval_max=interval_max,
            raw_zone=raw_zone,
            range_status=row["parse_status"],
            callable_interval_ids=callable_ids,
        )
    return None


def _expected(
    *,
    gap: Optional[float],
    low_side: str,
    interval: Optional[int],
    water: Optional[str],
    source: str,
    confidence: str,
    rule: str,
    notes: list[str],
    strength: StrengthContext | None = None,
    matched_row_id: str | None = None,
    matched_row_number: int | None = None,
    matched_sheet: str | None = None,
    lookup_key_status: str = "NO_TABLE_MATCH",
    interval_min: int | None = None,
    interval_max: int | None = None,
    raw_zone: str | None = None,
    range_status: str = "SINGLE_INTERVAL",
    callable_interval_ids: list[int] | None = None,
) -> ExpectedOpeningInterval:
    key = p4_strength_key(gap)
    if interval is not None:
        interval_min = interval if interval_min is None else interval_min
        interval_max = interval if interval_max is None else interval_max
    callable_ids = callable_interval_ids or ([] if interval is None else [interval])
    interval_text = None
    if interval is not None:
        if interval_min is not None and interval_max is not None and interval_min != interval_max:
            interval_text = f"{low_side} / {interval_min}-{interval_max}\u533a / {water}"
        else:
            interval_text = f"{low_side} / {interval}\u533a / {water}"
    return ExpectedOpeningInterval(
        final_gap_value=gap,
        final_gap_label=_gap_label(gap),
        expected_low_side=low_side,
        expected_interval_id=interval,
        expected_water_band=water,
        source=source,
        confidence=confidence,
        rule=rule,
        notes=notes,
        expected_interval_min_id=interval_min,
        expected_interval_max_id=interval_max,
        expected_interval_raw_zone=raw_zone,
        expected_interval_range_status=range_status,
        callable_interval_ids=callable_ids,
        p4_strength_key=key,
        static_strength_gap=getattr(strength, "static_gap", None),
        dynamic_adjustment=getattr(strength, "dynamic_adjustment", None),
        final_dynamic_strength_relation=getattr(strength, "final_gap", None),
        expected_interval_source=source,
        matched_sheet=matched_sheet or (P4_BRIDGE_SHEET if source == "STRENGTH_INTERVAL_BRIDGE" else None),
        matched_row_id=matched_row_id,
        matched_row_number=matched_row_number,
        lookup_key_status=lookup_key_status,
        expected_interval=interval_text,
        expected_interval_confidence=confidence,
    )


def expected_interval_from_gap(gap: Optional[float]) -> ExpectedOpeningInterval:
    """Debug/test fallback only. Formal analysis must use the P4 bridge table."""
    notes = ["RULE_FALLBACK 仅允许用于调试或测试，不得用于正式结构输出。"]
    if gap is None:
        return _expected(
            gap=None,
            low_side="未解析",
            interval=None,
            water=None,
            source="REVIEW_REQUIRED",
            confidence="REVIEW_REQUIRED",
            rule="动态广义实力关系无法解析。",
            notes=notes,
        )
    if gap >= 1.75:
        interval, low_side, water = min(10, 2 + int(round(gap / 0.5))), "主低赔", "中水"
    elif gap >= 1.25:
        interval, low_side, water = 2 + int(round(gap / 0.5)), "主低赔", "中水"
    elif gap >= 0.75:
        interval, low_side, water = 2 + int(round(gap / 0.5)), "主低赔", "中水"
    elif gap >= 0.25:
        interval, low_side, water = 3, "主低赔", "中水"
    elif gap > -0.25:
        interval, low_side, water = 2, "主低赔", "中水"
    elif gap >= -1.0:
        interval, low_side, water = max(0, 2 + int(round(gap / 0.5))), "客低赔", "中水"
        notes.append("客队高0.5-1档时，fallback 按每0.5档移动一个骨架区间。")
    else:
        interval, low_side, water = 0, "客低赔", "高水"
    return _expected(
        gap=gap,
        low_side=low_side,
        interval=interval,
        water=water,
        source="RULE_FALLBACK",
        confidence="DEBUG_ONLY",
        rule="档位差 fallback 只用于调试或测试。",
        notes=notes,
        lookup_key_status="RULE_FALLBACK",
    )


def expected_interval_from_table(
    *,
    strength: StrengthContext,
    estimate=None,
    bridge_path: str | Path | None = None,
    table_path: str | Path | None = None,
) -> ExpectedOpeningInterval:
    """Read the formal P4-to-modern-skeleton bridge table."""
    gap = parse_gap_value(strength, estimate)
    v6_expected = _expected_interval_from_v6_grade_map(table_path=table_path, strength=strength, gap=gap)
    if v6_expected is not None:
        return v6_expected
    key = p4_strength_key(gap)
    if key is None:
        return _expected(
            gap=None,
            low_side="未解析",
            interval=None,
            water=None,
            source="REVIEW_REQUIRED",
            confidence="REVIEW_REQUIRED",
            rule="无法生成 p4_strength_key，禁止进入理论区间。",
            notes=["需要人工校准动态广义实力关系。"],
            strength=strength,
        )
    path = Path(bridge_path) if bridge_path else DEFAULT_P4_BRIDGE
    if not path.is_file():
        return _expected(
            gap=gap,
            low_side="未解析",
            interval=None,
            water=None,
            source="REVIEW_REQUIRED",
            confidence="REVIEW_REQUIRED",
            rule="P4→现代骨架桥接表不存在，禁止进入理论区间。",
            notes=[f"桥接表路径={path}"],
            strength=strength,
        )
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row_number, row in enumerate(csv.DictReader(handle), start=2):
            if row.get("p4_strength_key") != key:
                continue
            try:
                interval = int(str(row.get("expected_interval_id", "")).strip())
            except ValueError:
                interval = None
            low_side = str(row.get("expected_low_side") or "").strip()
            water = str(row.get("expected_water_band") or "").strip()
            status = str(row.get("lookup_status") or "").strip()
            if interval is None or not low_side or status != "TABLE_CALLABLE":
                break
            return _expected(
                gap=gap,
                low_side=low_side,
                interval=interval,
                water=water or None,
                source="STRENGTH_INTERVAL_BRIDGE",
                confidence=str(row.get("expected_interval_confidence") or "TABLE_CONFIRMED"),
                rule="P4 动态广义实力键已通过桥接表连接现代骨架理论区间。",
                notes=[str(row.get("reasoning") or "表内桥接完成。")],
                strength=strength,
                matched_row_id=str(row.get("row_id") or ""),
                matched_row_number=row_number,
                lookup_key_status="TABLE_MATCHED",
            )
    return _expected(
        gap=gap,
        low_side="未解析",
        interval=None,
        water=None,
        source="REVIEW_REQUIRED",
        confidence="REVIEW_REQUIRED",
        rule="p4_strength_key 无表内匹配，禁止进入理论区间。",
        notes=[f"未匹配键={key}"],
        strength=strength,
        lookup_key_status="NO_TABLE_MATCH",
    )


def _opening_low_coordinate(company_set):
    for coordinate in company_set.coordinates:
        if coordinate.time_point == "initial" and coordinate.is_snapshot_low:
            return coordinate
    return None


def _semantic_tags(coord, expected: ExpectedOpeningInterval, delta: Optional[int]) -> tuple[list[str], str, str]:
    if coord is None or expected.expected_interval_id is None or delta is None:
        return [], "无法比较理论区间与现实初赔。", "UNCONFIRMED"
    gap = expected.final_gap_value
    if delta == 0:
        return ["顺区间"], "现实初赔低赔项与表驱动理论低赔区间一致。", "CONFIRMED"
    if coord.table_direction == "主低赔":
        if delta > 0:
            if gap is not None and gap < 0:
                return ["抬胜", "拉平", "负韬"], "客队广义实力更高，但现实初赔把主胜做入更深区间。", "CONFIRMED"
            return ["主胜偏实"], "现实初赔比理论区间更深，需结合三项组合解释。", "CONFIRMED"
        return ["胜韬", "平负合力备选"], "现实初赔比理论区间更浅，需检查平负组合。", "CONFIRMED"
    if coord.table_direction == "客低赔":
        if delta > 0:
            return ["抬负"], "现实初赔比理论区间更深，需区分客胜实盘与客向诱导。", "CONFIRMED"
        return ["负韬", "胜平合力备选"], "现实初赔比理论区间更浅，需检查胜平组合。", "CONFIRMED"
    return ["平低赔特殊"], "平赔成为低赔方向，必须进入平赔专项解释。", "SPECIAL_DRAW_LOW"


def _range_deviation(value: Optional[float], lower: Optional[float], upper: Optional[float]) -> Optional[float]:
    if value is None or lower is None or upper is None:
        return None
    if value < lower:
        return round(value - lower, 6)
    if value > upper:
        return round(value - upper, 6)
    return 0.0


def _reasonableness(low_deviation: Optional[float], profile_status: str) -> str:
    if profile_status != "PROFILE_CONFIRMED" or low_deviation is None:
        return "REVIEW_REQUIRED"
    if low_deviation < 0:
        return "DEEPER_THAN_THEORETICAL_RANGE"
    if low_deviation > 0:
        return "SHALLOWER_THAN_THEORETICAL_RANGE"
    return "WITHIN_THEORETICAL_RANGE"


def _expected_bounds(expected: ExpectedOpeningInterval) -> tuple[Optional[int], Optional[int]]:
    lower = expected.expected_interval_min_id
    upper = expected.expected_interval_max_id
    if lower is None:
        lower = expected.expected_interval_id
    if upper is None:
        upper = expected.expected_interval_id
    return lower, upper


def _interval_delta(opening_interval: Optional[int], expected: ExpectedOpeningInterval) -> Optional[int]:
    if opening_interval is None:
        return None
    lower, upper = _expected_bounds(expected)
    if lower is None or upper is None:
        return None
    if lower <= opening_interval <= upper:
        return 0
    if opening_interval < lower:
        return opening_interval - lower
    return opening_interval - upper


def _profile_for_expected_range(
    *,
    xlsx_path: str | Path | None,
    system: str,
    expected: ExpectedOpeningInterval,
) -> SkeletonIntervalProfile | None:
    lower, upper = _expected_bounds(expected)
    if xlsx_path is None or lower is None or upper is None:
        return None
    profiles = [load_interval_profile(str(xlsx_path), system, interval_id) for interval_id in range(lower, upper + 1)]
    if not profiles:
        return None
    main_values = [value for item in profiles for value in (item.main_price_min, item.main_price_max) if value is not None]
    draw_values = [value for item in profiles for value in (item.draw_reference_min, item.draw_reference_max) if value is not None]
    away_values = [value for item in profiles for value in (item.away_reference_min, item.away_reference_max) if value is not None]
    status = "PROFILE_CONFIRMED" if main_values and all(item.status == "PROFILE_CONFIRMED" for item in profiles) else "PROFILE_REVIEW_REQUIRED"
    notes = [note for item in profiles for note in item.notes]
    if lower != upper:
        notes.append(f"EXPECTED_INTERVAL_RANGE: using combined skeleton profile for {lower}-{upper} intervals.")
    return SkeletonIntervalProfile(
        system=system,
        sheet_name=profiles[0].sheet_name,
        interval_id=lower,
        main_price_min=min(main_values) if main_values else None,
        main_price_max=max(main_values) if main_values else None,
        draw_reference_min=min(draw_values) if draw_values else None,
        draw_reference_max=max(draw_values) if draw_values else None,
        away_reference_min=min(away_values) if away_values else None,
        away_reference_max=max(away_values) if away_values else None,
        row_numbers=[row for item in profiles for row in item.row_numbers],
        status=status,
        notes=notes,
    )


def audit_opening_interval(
    *,
    strength: StrengthContext,
    estimate,
    odds_coordinates,
    bridge_path: str | Path | None = None,
    allow_rule_fallback: bool = False,
    expected: ExpectedOpeningInterval | None = None,
    xlsx_path: str | Path | None = None,
) -> IntervalAuditResult:
    expected = expected or expected_interval_from_table(
        strength=strength,
        estimate=estimate,
        bridge_path=bridge_path,
        table_path=xlsx_path,
    )
    if expected.expected_interval_source != "STRENGTH_INTERVAL_BRIDGE" and allow_rule_fallback:
        expected = expected_interval_from_gap(parse_gap_value(strength, estimate))
    result = IntervalAuditResult(expected=expected, audits=[], ok=False, notes=[])
    if expected.expected_interval_source != "STRENGTH_INTERVAL_BRIDGE":
        result.notes.append(
            f"EXPECTED_INTERVAL_STATUS = REVIEW_REQUIRED｜expected_interval_source = {expected.expected_interval_source}。"
        )
        return result
    if odds_coordinates is None:
        result.stop_reason = "未生成赔率坐标，不能比较理论区间与机构初赔。"
        return result

    hard_seen = 0
    for company_set in odds_coordinates.company_sets:
        company_cn = _company_cn(company_set.company)
        if company_cn not in {"威廉", "立博"}:
            continue
        hard_seen += 1
        coord = _opening_low_coordinate(company_set)
        if coord is None:
            result.audits.append(OpeningIntervalAudit(
                company=company_cn,
                opening_low_direction="未解析",
                opening_low_odds=0.0,
                opening_interval_id=None,
                opening_water_band=None,
                expected_low_side=expected.expected_low_side,
                expected_interval_id=expected.expected_interval_id,
                expected_interval_min_id=expected.expected_interval_min_id,
                expected_interval_max_id=expected.expected_interval_max_id,
                expected_interval_raw_zone=expected.expected_interval_raw_zone,
                interval_delta=None,
                deviation_label="缺少初赔低赔坐标",
                interpretation="硬判断公司初赔低赔方向未能坐标归位。",
                hard_status="NO_OPENING_COORDINATE",
            ))
            continue
        delta = _interval_delta(coord.interval_id, expected)
        tags, interpretation, hard_status = _semantic_tags(coord, expected, delta)
        profile = _profile_for_expected_range(xlsx_path=xlsx_path, system=coord.system, expected=expected)
        profile_status = profile.status if profile else "PROFILE_REVIEW_REQUIRED"
        low_deviation = _range_deviation(
            float(coord.actual_low_odds),
            getattr(profile, "main_price_min", None),
            getattr(profile, "main_price_max", None),
        )
        draw_deviation = _range_deviation(
            float(coord.odds_draw),
            getattr(profile, "draw_reference_min", None),
            getattr(profile, "draw_reference_max", None),
        )
        away_deviation = _range_deviation(
            float(coord.odds_away),
            getattr(profile, "away_reference_min", None),
            getattr(profile, "away_reference_max", None),
        )
        price_reasonableness = _reasonableness(low_deviation, profile_status)
        if profile_status != "PROFILE_CONFIRMED":
            hard_status = "EXPECTED_SKELETON_REVIEW_REQUIRED"
            interpretation = (
                f"{coord.system} 的理论 {expected.expected_interval_id} 区没有可调用低赔精确骨架，"
                "不得推断机构初赔动机。"
            )
        expected_zone = expected.expected_interval_raw_zone or (
            f"{expected.expected_interval_min_id}-{expected.expected_interval_max_id}区"
            if expected.expected_interval_min_id is not None
            and expected.expected_interval_max_id is not None
            and expected.expected_interval_min_id != expected.expected_interval_max_id
            else f"{expected.expected_interval_id}区"
        )
        deviation = (
            "无法比较"
            if delta is None
            else f"落入理论范围{expected_zone}"
            if delta == 0
            else f"现实深于理论范围{delta}区"
            if delta > 0
            else f"现实浅于理论范围{abs(delta)}区"
        )
        result.audits.append(OpeningIntervalAudit(
            company=company_cn,
            opening_low_direction=LOW_SIDE_TO_DIRECTION.get(coord.table_direction, coord.direction),
            opening_low_odds=float(coord.actual_low_odds),
            opening_interval_id=coord.interval_id,
            opening_water_band=coord.water_band,
            expected_low_side=expected.expected_low_side,
            expected_interval_id=expected.expected_interval_id,
            expected_interval_min_id=expected.expected_interval_min_id,
            expected_interval_max_id=expected.expected_interval_max_id,
            expected_interval_raw_zone=expected.expected_interval_raw_zone,
            interval_delta=delta,
            deviation_label=deviation,
            semantic_tags=tags,
            interpretation=interpretation,
            hard_status=hard_status,
            system=coord.system,
            sheet_name=getattr(profile, "sheet_name", coord.sheet_name),
            raw_opening_home=float(coord.odds_home),
            raw_opening_draw=float(coord.odds_draw),
            raw_opening_away=float(coord.odds_away),
            # Backward-compatible aliases: numerical conversion is disabled.
            converted_opening_home=float(coord.odds_home),
            converted_opening_draw=float(coord.odds_draw),
            converted_opening_away=float(coord.odds_away),
            expected_home_min=getattr(profile, "main_price_min", None),
            expected_home_max=getattr(profile, "main_price_max", None),
            expected_draw_reference_min=getattr(profile, "draw_reference_min", None),
            expected_draw_reference_max=getattr(profile, "draw_reference_max", None),
            expected_away_reference_min=getattr(profile, "away_reference_min", None),
            expected_away_reference_max=getattr(profile, "away_reference_max", None),
            home_range_deviation=low_deviation,
            draw_reference_deviation=draw_deviation,
            away_reference_deviation=away_deviation,
            price_reasonableness=price_reasonableness,
            skeleton_profile_status=profile_status,
        ))

    if hard_seen < 2:
        result.stop_reason = "缺少 William / Ladbrokes 初赔坐标，不能完成表驱动理论区间审计。"
        return result
    if any(a.hard_status != "CONFIRMED" for a in result.audits):
        result.stop_reason = "William / Ladbrokes 初赔区间审计未全部完成，禁止进入变赔解释。"
        return result
    result.ok = True
    return result
