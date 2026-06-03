from __future__ import annotations

import re
from functools import lru_cache
from typing import Any, Optional

import openpyxl

from .models import OddsSystemConversion, SkeletonIntervalProfile, TableLookupResult
from .returns import low_direction, low_odds

SYSTEM_SHEETS = {f"{value}系": f"{value}体系" for value in range(89, 97)}
MAIN_PRICE_COLUMN = "主赔_骨架精确"
DRAW_REFERENCE_COLUMN = "平赔_机构档口参考"
AWAY_REFERENCE_COLUMN = "负赔_机构档口参考"


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(str(value).replace("%", "").strip())
    except ValueError:
        return None


def _interval(value: Any) -> Optional[int]:
    if value is None:
        return None
    match = re.search(r"\d+", str(value))
    return int(match.group()) if match else None


@lru_cache(maxsize=24)
def load_system_rows(xlsx_path: str, system: str) -> list[dict[str, Any]]:
    sheet_name = SYSTEM_SHEETS.get(system)
    if sheet_name is None:
        raise ValueError(f"OUT_OF_89_96_SYSTEM: {system}")
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"NO_SYSTEM_SHEET: {sheet_name}")
        sheet = workbook[sheet_name]
        rows_iter = sheet.iter_rows(values_only=True)
        next(rows_iter)  # title row
        headers = [str(value) if value is not None else "" for value in next(rows_iter)]
        rows: list[dict[str, Any]] = []
        for row_number, values in enumerate(rows_iter, start=3):
            item = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
            if _to_float(item.get(MAIN_PRICE_COLUMN)) is None:
                continue
            item["__row_number__"] = row_number
            rows.append(item)
        return rows
    finally:
        workbook.close()


def _row_score(row: dict[str, Any], actual_home_odds: float) -> float:
    return abs((_to_float(row.get(MAIN_PRICE_COLUMN)) or 999.0) - actual_home_odds)


def _bounds(rows: list[dict[str, Any]], selected: dict[str, Any]) -> tuple[float, float, float, str]:
    reference = _to_float(selected.get(MAIN_PRICE_COLUMN))
    if reference is None:
        raise ValueError("MAIN_PRICE_AXIS_MISSING")
    prices = sorted({_to_float(row.get(MAIN_PRICE_COLUMN)) for row in rows if _to_float(row.get(MAIN_PRICE_COLUMN)) is not None})
    index = prices.index(reference)
    previous_value = prices[index - 1] if index > 0 else reference - 0.10
    next_value = prices[index + 1] if index + 1 < len(prices) else reference + 0.10
    lower = round((previous_value + reference) / 2, 6)
    upper = round((reference + next_value) / 2, 6)
    return lower, upper, reference, str(selected.get("水位") or "")


def _range(values: list[Optional[float]]) -> tuple[Optional[float], Optional[float]]:
    numeric = [value for value in values if value is not None]
    return (min(numeric), max(numeric)) if numeric else (None, None)


def load_interval_profile(xlsx_path: str, system: str, interval_id: int) -> SkeletonIntervalProfile:
    """Read the bookmaker-system skeleton values for one theoretical interval."""

    rows = [row for row in load_system_rows(xlsx_path, system) if _interval(row.get("区间")) == interval_id]
    main_min, main_max = _range([_to_float(row.get(MAIN_PRICE_COLUMN)) for row in rows])
    draw_min, draw_max = _range([_to_float(row.get(DRAW_REFERENCE_COLUMN)) for row in rows])
    away_min, away_max = _range([_to_float(row.get(AWAY_REFERENCE_COLUMN)) for row in rows])
    status = "PROFILE_CONFIRMED" if main_min is not None and main_max is not None else "PROFILE_REVIEW_REQUIRED"
    notes = []
    if status != "PROFILE_CONFIRMED":
        notes.append(f"{system} {interval_id}区没有可调用的主赔精确骨架。")
    if draw_min is None or draw_max is None or away_min is None or away_max is None:
        notes.append("平赔或负赔机构档口参考不完整。")
    return SkeletonIntervalProfile(
        system=system,
        sheet_name=SYSTEM_SHEETS.get(system, system),
        interval_id=interval_id,
        main_price_min=main_min,
        main_price_max=main_max,
        draw_reference_min=draw_min,
        draw_reference_max=draw_max,
        away_reference_min=away_min,
        away_reference_max=away_max,
        row_numbers=[int(row["__row_number__"]) for row in rows],
        status=status,
        notes=notes,
    )


def lookup_company_odds(
    *,
    xlsx_path: str,
    league: str = "",
    company: str,
    conversion: OddsSystemConversion,
) -> TableLookupResult:
    """Read the corrected market-ladder workbook by detected return-rate system.

    `league` is retained only as context metadata. It never controls table access.
    """
    if not isinstance(conversion, OddsSystemConversion):
        raise ValueError("TABLE_LOOKUP_FORBIDDEN: table_lookup 只能消费 odds_system 体系路由结果。")
    if conversion.system_lookup_status != "SYSTEM_LOOKUP_ALLOWED":
        raise ValueError(f"TABLE_LOOKUP_FORBIDDEN: {conversion.system_lookup_status}")

    # Keep institution-published odds unchanged. Return-rate normalization is
    # represented by selecting the matching 89-96 system sheet.
    snapshot = conversion.comparison_snapshot()
    actual_low = low_odds(snapshot)
    system = conversion.target_system
    rows = load_system_rows(xlsx_path, system)
    if not rows:
        raise ValueError(f"NO_TABLE_ROW: {system}")
    selected = min(rows, key=lambda row: _row_score(row, snapshot.home))
    lower, upper, reference, water_band = _bounds(rows, selected)
    if snapshot.home < lower:
        deviation = "低于表内下界"
        boundary_distance = lower - snapshot.home
    elif snapshot.home > upper:
        deviation = "高于表内上界"
        boundary_distance = snapshot.home - upper
    else:
        deviation = "表内"
        boundary_distance = min(snapshot.home - lower, upper - snapshot.home)
    direction = low_direction(snapshot)
    return TableLookupResult(
        company=company,
        system=system,
        league=league,
        direction=direction,
        interval_id=_interval(selected.get("区间")),
        water_band=water_band,
        lower_bound=lower,
        upper_bound=upper,
        table_reference_odds=reference,
        actual_low_odds=float(actual_low),
        boundary_distance=round(boundary_distance, 6),
        deviation=deviation,
        sheet_name=SYSTEM_SHEETS[system],
        row_number=int(selected["__row_number__"]),
        lookup_status="TABLE_READ_CONFIRMED",
        table_axis="home",
        table_axis_odds=float(snapshot.home),
        raw_row=selected,
    )
